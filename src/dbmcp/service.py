"""核心服务层：与 MCP 传输解耦，便于单元测试。

所有会触达数据库的操作都必须落审计记录（成功 / 拒绝 / 出错），
拒绝路径同样入库——这正是要给人看的部分。
"""

from __future__ import annotations

import hmac
import logging
import threading
from dataclasses import dataclass
from typing import TYPE_CHECKING

from .approvals import ApprovalError, ApprovalStore
from .audit.classify import classify, fingerprint
from .audit.log import AuditRecord, AuditStore
from .audit.redis_rules import classify_command, command_fingerprint, parse_command
from .audit.risk import assess
from .config import AppConfig, ConnectionConfig
from .health import ConnectionUnavailable, HealthMonitor, is_connection_error
from .masking import apply_mask
from .metadata import MetadataCache
from .notify import NoopNotifier, Notifier
from . import engines, redis_engine

if TYPE_CHECKING:
    from .snippets import SnippetStore

logger = logging.getLogger(__name__)

HOUSEKEEPING_INTERVAL_S = 60
DEFAULT_RETENTION_DAYS = 30
ADMIN_PAGE_SIZE = 100  # 查询台每页行数（上限受连接 max_rows 约束）
DEFAULT_AGENT_MAX_RESULT_CHARS = 40000  # agent 结果字符预算的最终兜底（settings 未启用时）


class QueryRejected(Exception):
    """SQL 被审计规则拒绝。message 面向 agent，说明原因与下一步动作。"""


def _is_no_database_error(e: Exception) -> bool:
    """识别"未选定数据库"类错误：MySQL 1046 / PG no schema / 未限定表名。"""
    msg = str(e).lower()
    return (
        "1046" in msg
        or "no database selected" in msg
        or "no schema has been selected" in msg
    )


def _rows_to_text(columns: list[str], rows: list[list], max_rows: int = 5) -> str:
    """把样本行拼成紧凑 TSV（喂 AI 用）：首行列名，其余数据行，制表符分隔，None→\\N。"""
    def cell(v: object) -> str:
        if v is None:
            return "\\N"
        return str(v).replace("\t", " ").replace("\n", " ")
    lines = ["\t".join(columns)]
    for row in rows[:max_rows]:
        lines.append("\t".join(cell(v) for v in row))
    return "\n".join(lines)


def _ai_api_cfg(s: dict) -> dict:
    """从设置里取 provider=api 的连接配置（base/format/key_env）。"""
    return {"base": str(s.get("ai_api_base") or ""),
            "format": str(s.get("ai_api_format") or "anthropic"),
            "key_env": str(s.get("ai_api_key_env") or "")}


def _layout_graph(graph: dict) -> None:
    """给 AI 生成的节点按拓扑层级赋 x/y（AI 不给坐标），使画布排版可读。原地修改。"""
    nodes = graph.get("nodes") or []
    ids = {n.get("id") for n in nodes}
    preds: dict = {n.get("id"): [] for n in nodes}
    for e in graph.get("edges") or []:
        if e.get("from") in ids and e.get("to") in ids:
            preds[e["to"]].append(e["from"])
    level: dict = {}

    def _lvl(nid: str, seen: frozenset) -> int:
        if nid in level:
            return level[nid]
        ps = [p for p in preds.get(nid, []) if p not in seen]
        level[nid] = 0 if not ps else 1 + max(_lvl(p, seen | {nid}) for p in ps)
        return level[nid]

    for n in nodes:
        _lvl(n.get("id"), frozenset())
    per_level: dict = {}
    for n in nodes:
        lv = level.get(n.get("id"), 0)
        row = per_level.get(lv, 0)
        per_level[lv] = row + 1
        n["x"] = 30 + lv * 200
        n["y"] = 30 + row * 100


def _plan_node_name(plan: dict, node_id: str, graph: dict) -> str | None:
    """从编译后的 plan 找目标节点的名字（=工作区里的表/视图名）。

    plan 里 sources 有 node/dataset，steps 有 node/name；这些都是 compile_graph 输出。
    output 节点是虚 output_sql 无独立 view，此时回退取 graph.nodes[node_id].name。
    """
    for src in plan.get("sources") or []:
        if src.get("node") == node_id:
            return src.get("dataset")
    for st in plan.get("steps") or []:
        if st.get("node") == node_id:
            return st.get("name")
    for n in graph.get("nodes") or []:
        if n.get("id") == node_id:
            return (n.get("name") or "").strip() or None
    return None


def _preview_target_for(graph: dict, node_id: str) -> tuple[str, str] | None:
    """预览节点输出时应该看哪张 view/table。

    普通节点（source/file/filter/join/aggregate/sql）compile 时会 CREATE OR REPLACE
    VIEW <节点名>，DESCRIBE 节点名即可拿列。
    **output 节点例外**：SQL 是 `SELECT * FROM 上游 ORDER BY .. LIMIT ..`，不物化
    成 view；查询 output.name 会报 "Table … does not exist"。预览它=预览它的上游。

    返回 (target_view_name, source_node_id_to_materialize)：
    - 普通节点：(name, node_id)
    - output 节点：(上游节点的 name, 上游 node_id)
    找不到节点或 output 无上游连线 → None。
    """
    nodes = {n.get("id"): n for n in (graph.get("nodes") or [])}
    node = nodes.get(node_id)
    if not node:
        return None
    if node.get("type") != "output":
        name = (node.get("name") or "").strip()
        return (name, node_id) if name else None
    # output 节点：找 in 边指向它的那个上游
    for e in graph.get("edges") or []:
        if e.get("to") == node_id:
            up_id = e.get("from")
            up = nodes.get(up_id)
            if up:
                up_name = (up.get("name") or "").strip()
                if up_name:
                    return (up_name, up_id)
    return None


def _plan_prefix_for(plan: dict, target_name: str) -> dict:
    """从 plan 里挑出「构建 target_name 所必需的」sources + steps（按 plan 顺序，保拓扑序）。

    简单实现：走 SQL 里的 FROM/JOIN 关联反向追依赖太脆；这里保守起见——
    截到 target 在 steps 中出现的位置为止（含），sources 全带上（不多也不害事）。
    """
    sources = list(plan.get("sources") or [])
    steps: list[dict] = []
    for st in plan.get("steps") or []:
        steps.append(st)
        if st.get("name") == target_name:
            break
    return {"sources": sources, "steps": steps}


def _dataset_exists(store, workspace: str, name: str) -> bool:
    """工作区中是否已存在同名 table 或 view。"""
    try:
        con = store._connect(workspace, must_exist=False)
    except Exception:  # noqa: BLE001
        return False
    try:
        row = con.execute(
            "SELECT 1 FROM information_schema.tables"
            " WHERE table_schema='main' AND table_name = ? LIMIT 1",
            [name]).fetchone()
        return row is not None
    finally:
        con.close()


def _describe_columns(store, workspace: str, name: str) -> list[dict]:
    """DuckDB DESCRIBE 拿列名/类型。"""
    con = store._connect(workspace)
    try:
        rows = con.execute(f'DESCRIBE "{name}"').fetchall()
    finally:
        con.close()
    # DESCRIBE 返回 (column_name, column_type, null, key, default, extra)
    return [{"name": r[0], "type": r[1]} for r in rows]


@dataclass
class CallerInfo:
    agent: str = "unknown"
    session_id: str = ""


class DbmService:
    def __init__(
        self,
        config: AppConfig,
        store: AuditStore,
        approvals: ApprovalStore | None = None,
        metadata: MetadataCache | None = None,
        config_path: str | None = None,
        snippets: "SnippetStore | None" = None,
        notifier: Notifier | None = None,
    ):
        self.config = config
        self.store = store
        self.pool = engines.EnginePool()
        self.redis_pool = redis_engine.RedisPool()
        # 让引擎池能解析每跳的 SSH 证书引用（同一 dict，连接管理原地增删即时可见）
        self.pool.identities = self.config.ssh_identities
        self.redis_pool.identities = self.config.ssh_identities
        self.approvals = approvals
        # 通知抽象：默认 Noop（safe default，测试与库使用都不会真发通知）；
        # serve 入口注入 NotifierRouter（内推 + 用户配置的外部渠道，动态跟随设置）。
        self.notifier = notifier if notifier is not None else NoopNotifier()
        # 站内通知收件箱（serve 时注入 InboxStore）；SSE 铃铛与外部渠道之外的默认路径
        self.inbox = None
        # 健康监控：exhausted 时发通知（同一连接短时间去重）
        self.health = HealthMonitor(
            probe=self._health_probe,
            on_exhausted=self._on_connection_exhausted,
        )
        self.metadata = metadata
        self.config_path = config_path
        self.snippets = snippets
        self.settings = None   # SettingsStore（serve 时注入）
        self.analysis = None   # AnalysisStore（serve 时注入；未启用则分析功能不可用）
        self.workflows = None  # WorkflowStore（serve 时注入）
        self.schedules = None  # WorkflowScheduleStore（serve 时注入）
        self.runs = None       # WorkflowRunStore（serve 时注入）
        self._housekeeping_stop: threading.Event | None = None
        self._scheduler_stop: threading.Event | None = None
        # 记录本进程周期内已经为哪一分钟执行过（防止 30s tick 在同一分钟触发两次）
        self._sched_ticked_minute: set[tuple[str, str]] = set()  # {(name, "YYYY-MM-DD HH:MM")}
        self.data_dir = None   # serve 时注入，供 xlsx 产物落盘
        self.base_url = ""     # serve 时注入，如 http://127.0.0.1:8100（导出下载链接）

    # ---------- 元信息 ----------

    def list_projects(self) -> list[dict]:
        # 对 agent 隐藏 Redis 连接（Redis 只供人通过 /admin/redis 操作）；
        # 只剩 Redis 连接的项目也不出现
        out = []
        for name, proj in sorted(self.config.projects.items()):
            conns = sorted(n for n, c in proj.connections.items() if c.engine != "redis")
            if conns:
                out.append({"project": name, "connections": conns})
        return out

    def list_connections(self, project: str) -> list[dict]:
        proj = self.config.projects.get(project)
        if proj is None:
            raise KeyError(f"项目 {project!r} 不存在")
        return [
            {
                "connection": name,
                "engine": c.engine,
                "environment": c.environment,
                "database": c.database,
                "host": c.host,
                # 无默认库时提示 agent 用全限定表名
                **({"note": "此连接未绑定默认库，查询/schema 操作请用「库名.表名」全限定，"
                            "list_tables/describe_table 需先用 SHOW DATABASES 选定库"}
                   if c.engine in ("mysql", "postgres", "clickhouse") and not c.database else {}),
                # 有意不返回 user/password/writer 等账号信息
            }
            # Redis 有意不返回：agent 碰不到 Redis
            for name, c in sorted(proj.connections.items()) if c.engine != "redis"
        ]

    # ---------- 查询 ----------

    def _read(
        self, project: str, connection: str, cfg: ConnectionConfig, sql: str,
        caller: CallerInfo, max_rows: int, schema: str | None = None,
        on_start=None,  # noqa: ANN001
        max_cell_chars: int | None = None, mask: bool = True,
    ) -> dict:
        """执行一条已判定只读的 SQL：跑 reader、落审计、脱敏，返回结果 dict。

        max_rows 由调用方决定（query 用连接策略；查询台分页用 page_size+1 以探测下一页），
        与 truncated 检测解耦，便于复用。schema 为查询台的执行 schema 上下文。
        max_cell_chars 缺省用连接策略（查询台可传系统设置的 sql_max_cell_chars 覆盖）。
        mask=True 对敏感列脱敏（agent 路径的红线：密码不出现在工具返回值中）；已认证的后台
        查询台/导出传 mask=False——人就是要看真实数据，脱敏反而碍事。
        """
        rec = self._base_record(project, connection, cfg, "query", sql, caller)
        if schema:
            rec.detail = f"schema={schema}"

        def _do() -> "engines.QueryResult":
            engine = self.pool.get(project, connection, cfg, schema=schema)
            return engines.run_query(
                engine, sql, max_rows,
                max_cell_chars=max_cell_chars or cfg.policy.max_cell_chars,
                on_start=on_start,
            )

        try:
            result = self._run_touching_db(project, connection, _do)
        except ConnectionUnavailable as e:
            rec.status = "error"
            rec.detail = f"ConnectionUnavailable[{e.state}]: {e}"
            self.store.record(rec)
            raise
        except QueryRejected:
            raise
        except Exception as e:
            if not cfg.database and _is_no_database_error(e):
                rec.status = "error"
                rec.detail = "未选定数据库"
                self.store.record(rec)
                raise QueryRejected(
                    "该连接未绑定默认库。请用「库名.表名」全限定表名查询"
                    "（如 SELECT * FROM mydb.users），或先执行 SHOW DATABASES 查看可用库。"
                ) from e
            rec.status = "error"
            rec.detail = f"{type(e).__name__}: {e}"
            self.store.record(rec)
            raise

        rec.status = "ok"
        rec.row_count = result.row_count
        rec.duration_ms = result.duration_ms
        self.store.record(rec)
        rows, masked = apply_mask(result.columns, result.rows, cfg.policy) if mask else (result.rows, [])
        out = {
            "columns": result.columns,
            "rows": rows,
            "row_count": result.row_count,
            "truncated": result.truncated,
            "duration_ms": result.duration_ms,
            "column_types": result.column_types,
        }
        if masked:
            out["masked_columns"] = masked
        return out

    def query(self, project: str, connection: str, sql: str, caller: CallerInfo) -> dict:
        cfg = self.config.get_connection(project, connection)
        verdict = classify(sql, cfg.engine)
        if not verdict.readonly:
            rec = self._base_record(project, connection, cfg, "query", sql, caller)
            rec.status = "rejected"
            rec.detail = verdict.reason
            self.store.record(rec)
            raise QueryRejected(
                f"已拒绝：{verdict.reason}。query 工具仅允许只读语句；"
                "数据变更操作需人工授权的 execute 流程（M3 上线后提供）。"
            )

        # 兜底：缺 LIMIT 的 SELECT 注入 LIMIT max_rows+1，防大表全量缓冲把 DB/进程拖挂
        run_sql, _, _ = engines.paginate_sql(sql, cfg.engine, cfg.policy.max_rows + 1, 0)
        out = self._read(project, connection, cfg, run_sql, caller, cfg.policy.max_rows)
        out["statement_kind"] = verdict.statement_kind
        if out["truncated"]:
            out["hint"] = (
                f"结果已截断到 {cfg.policy.max_rows} 行（连接策略 max_rows）。"
                "如需后续数据，请在 SQL 中用 LIMIT/OFFSET（或 WHERE 条件缩小范围）自行分页。"
            )
        return out

    # ---------- 管理后台查询台（人已认证，写操作二次确认后直接执行）----------

    def admin_run_sql(
        self, project: str, connection: str, sql: str, caller: CallerInfo, confirm: bool = False,
        page: int = 0, page_size: int | None = None, schema: str | None = None,
        on_start=None,  # noqa: ANN001
        confirm_text: str | None = None, expect_fingerprint: str | None = None,
    ) -> dict:
        """管理后台查询台专用入口。**只挂在已认证的后台路由上，agent 无法触达。**

        - 只读语句：跑 reader 出结果，自动分页（缺 LIMIT 的 SELECT 注入 LIMIT/OFFSET
          兜底，防大表拉挂 DB）；用户自带 LIMIT 则尊重不改。
        - 写语句 + confirm=False：评估风险并返回风险报告（含 fingerprint / prod / expect_text），
          **不执行**。
        - 写语句 + confirm=True：经人工二次确认，直接用 writer 账号执行并落审计。
          这是后台专属旁路（不进审批单）；红线「拒绝—重提」只约束 agent 的 execute。
          二次闸门 H1 指纹绑定：确认时若带回 expect_fingerprint，须与当前 SQL 的指纹一致，
          否则拒绝——防「看 A 批 B」（确认前后 SQL 被改）。
          注：prod 写操作只需人工二次确认（不再要求输入连接名），便利优先；红框/红条视觉警示仍在。
        - schema：执行 schema 上下文（右上角选择），未限定表名的 SQL 在该库下执行。
        """
        cfg = self.config.get_connection(project, connection)
        verdict = classify(sql, cfg.engine)
        # 语法错误：明确报语法错，不走"确认写操作"流程（默认拒绝仍成立——不执行）
        if verdict.statement_kind == "ParseError":
            rec = self._base_record(project, connection, cfg, "query", sql, caller)
            rec.status = "rejected"
            rec.detail = verdict.reason
            self.store.record(rec)
            return {"kind": "error", "error": f"SQL 语法错误：{verdict.reason.replace('SQL 解析失败: ', '')}"}
        if verdict.readonly:
            page = max(page, 0)
            default_size = int(self._setting("sql_page_size") or ADMIN_PAGE_SIZE)
            eff_cell = int(self._setting("sql_max_cell_chars") or cfg.policy.max_cell_chars)
            # 分页每页行数仍受连接策略上限（连接可显式限行）；单元格上限用系统设置
            size = min(page_size or default_size, cfg.policy.max_rows)
            paged_sql, paginated, ordered = engines.paginate_sql(
                sql, cfg.engine, size + 1, page * size)
            if paginated:
                # 取 size+1 行探测是否有下一页；不受连接 max_rows 二次截断影响
                out = self._read(project, connection, cfg, paged_sql, caller, size + 1,
                                 schema=schema, on_start=on_start, max_cell_chars=eff_cell, mask=False)
                rows = out["rows"]
                out["has_next"] = len(rows) > size
                out["rows"] = rows[:size]
                out["row_count"] = len(out["rows"])
                out.update(paginated=True, page=page, page_size=size, ordered=ordered)
                out.pop("truncated", None)
                return {"kind": "read", **out}
            # 自带 LIMIT / 非 SELECT：不分页，受系统设置的结果行上限 sql_max_rows 兜底
            eff_max_rows = int(self._setting("sql_max_rows") or cfg.policy.max_rows)
            out = self._read(project, connection, cfg, sql, caller, eff_max_rows,
                             schema=schema, on_start=on_start, max_cell_chars=eff_cell, mask=False)
            out["paginated"] = False
            return {"kind": "read", **out}

        is_prod = (cfg.environment or "").lower() == "prod"
        fp = fingerprint(sql, cfg.engine)
        if not confirm:
            report = assess(sql, cfg.engine, self._meta_provider(project, connection, cfg))
            report_dict = report.to_dict()
            plan = self._try_explain(project, connection, cfg, sql, schema=schema)
            if plan:
                report_dict["explain"] = plan
            return {"kind": "confirm", "risk": report_dict,
                    "statement_kind": verdict.statement_kind,
                    "fingerprint": fp, "prod": is_prod,
                    "expect_text": connection if is_prod else None}

        # H1：确认必须绑定到刚才被评估/展示的那条 SQL（指纹一致），否则拒绝执行
        if expect_fingerprint is not None and not hmac.compare_digest(expect_fingerprint, fp):
            rec = self._base_record(project, connection, cfg, "admin_execute", sql, caller)
            rec.status = "rejected"
            rec.detail = "确认指纹与提交 SQL 不一致，已拒绝执行（H1）"
            self.store.record(rec)
            raise QueryRejected("SQL 在确认前后发生了变化（指纹不一致），已拒绝执行，请重新确认。")

        rec = self._base_record(project, connection, cfg, "admin_execute", sql, caller)
        if schema:
            rec.detail = f"schema={schema}"

        def _do() -> "engines.QueryResult":
            engine = self.pool.get(project, connection, cfg, role="writer", schema=schema)
            return engines.run_write(engine, sql, on_start=on_start)

        try:
            result = self._run_touching_db(project, connection, _do)
        except ConnectionUnavailable as e:
            rec.status = "error"
            rec.detail = f"ConnectionUnavailable[{e.state}]: {e}"
            self.store.record(rec)
            raise
        except Exception as e:
            rec.status = "error"
            rec.detail = f"{type(e).__name__}: {e}"
            self.store.record(rec)
            raise
        rec.status = "ok"
        rec.detail = "后台查询台直接执行（已二次确认）" + (f" schema={schema}" if schema else "")
        rec.row_count = result.row_count
        rec.duration_ms = result.duration_ms
        self.store.record(rec)
        return {"kind": "write", "affected_rows": result.row_count,
                "duration_ms": result.duration_ms}

    MAX_IMPORT_ROWS = 50_000

    def admin_import_rows(
        self, project: str, connection: str, table: str, columns: list[str],
        rows: list[list], caller: CallerInfo, schema: str | None = None,
    ) -> dict:
        """后台数据导入（CSV/粘贴）：参数化批量 INSERT，writer 单事务执行并审计。

        安全：列名必须存在于目标表结构（防拼接注入）；值全部走绑定参数；
        行数上限 MAX_IMPORT_ROWS。仅挂在已认证的后台路由上，agent 无法触达。
        """
        cfg = self.config.get_connection(project, connection)
        if not rows:
            raise ValueError("没有可导入的行")
        if len(rows) > self.MAX_IMPORT_ROWS:
            raise ValueError(f"单次导入上限 {self.MAX_IMPORT_ROWS} 行，实际 {len(rows)} 行")
        if not columns:
            raise ValueError("缺少列映射")
        info = self.describe_table(project, connection, table, caller, schema=schema)
        valid = {c["name"] for c in info["columns"]}
        bad = [c for c in columns if c not in valid]
        if bad:
            raise ValueError(f"列不存在于表 {table}: {', '.join(bad)}（表列: {', '.join(sorted(valid))}）")
        rec = self._base_record(project, connection, cfg, "admin_import",
                                f"IMPORT INTO {table} ({', '.join(columns)}) — {len(rows)} 行",
                                caller)

        def _do() -> "engines.QueryResult":
            engine = self.pool.get(project, connection, cfg, role="writer", schema=schema)
            return engines.insert_rows(engine, table, columns, rows, schema=schema)

        try:
            result = self._run_touching_db(project, connection, _do)
        except ConnectionUnavailable as e:
            rec.status = "error"
            rec.detail = f"ConnectionUnavailable[{e.state}]: {e}"
            self.store.record(rec)
            raise
        except Exception as e:
            rec.status = "error"
            rec.detail = f"{type(e).__name__}: {e}"
            self.store.record(rec)
            raise
        rec.status = "ok"
        rec.row_count = result.row_count
        rec.duration_ms = result.duration_ms
        rec.detail = "后台导入（已确认，单事务）" + (f" schema={schema}" if schema else "")
        self.store.record(rec)
        return {"inserted": result.row_count, "duration_ms": result.duration_ms}

    def admin_export(
        self, project: str, connection: str, sql: str, fmt: str, caller: CallerInfo,
        schema: str | None = None,
    ) -> tuple[bytes, str, str]:
        """导出只读查询结果为文件，返回 (字节, media_type, 扩展名)。仅限只读语句。"""
        from .export import export_result

        cfg = self.config.get_connection(project, connection)
        if not classify(sql, cfg.engine).readonly:
            raise QueryRejected("导出仅支持只读查询（SELECT/SHOW/...）的结果")
        run_sql, _, _ = engines.paginate_sql(sql, cfg.engine, cfg.policy.max_rows + 1, 0)
        result = self._read(project, connection, cfg, run_sql, caller,
                            cfg.policy.max_rows, schema=schema, mask=False)
        return export_result(result["columns"], result["rows"], fmt)

    def export_table(
        self,
        project: str,
        connection: str,
        table: str,
        fields: list[str] | None,
        limit: int,
        fmt: str,
        caller: CallerInfo,
        database: str | None = None,
    ) -> dict:
        """供 agent 导出单表数据，文件落服务端，仅返回下载链接及摘要。

        表、库和字段均先通过数据库反射校验，再由当前方言引用标识符，不接受任意 SQL，
        从而避免把导出入口变成查询分类器的旁路。agent 导出沿用敏感字段脱敏策略；
        文件内容不进入 MCP tool result，避免大文件占用模型上下文。
        """
        from .export import SUPPORTED_FORMATS, export_result

        cfg = self.config.get_connection(project, connection)
        if fmt not in SUPPORTED_FORMATS:
            raise ValueError(f"不支持的导出格式 {fmt!r}，可选：{', '.join(SUPPORTED_FORMATS)}")
        if limit < 1:
            raise ValueError("导出行数必须大于 0")
        if limit > cfg.policy.max_rows:
            raise ValueError(
                f"导出行数 {limit} 超过连接策略上限 {cfg.policy.max_rows}，"
                "请减少行数或由管理员调整 max_rows"
            )

        # 兼容 table="库.表"；同时传 database 时要求两者一致，避免含糊选择。
        if "." in table:
            table_database, plain_table = table.split(".", 1)
            if database is not None and database != table_database:
                raise ValueError(
                    f"表名中的库 {table_database!r} 与 database={database!r} 不一致"
                )
            database, table = table_database, plain_table
        if not table:
            raise ValueError("表名不能为空")
        if database is None and not cfg.database and cfg.engine in (
            "mysql", "postgres", "clickhouse"
        ):
            raise ValueError("此连接未绑定默认库，请通过 database 参数选择要导出的库（schema）")

        engine = self.pool.get(project, connection, cfg, schema=database)
        info = engines.describe_table(engine, table, database)
        available = [str(c["name"]) for c in info["columns"]]
        selected = fields or available
        if not selected:
            raise ValueError(f"表 {table!r} 没有可导出的字段")
        if len(selected) != len(set(selected)):
            raise ValueError("导出字段不能重复")
        unknown = [name for name in selected if name not in available]
        if unknown:
            raise ValueError(
                f"字段不存在于表 {table}: {', '.join(unknown)}"
                f"（可选：{', '.join(available)}）"
            )

        preparer = engine.dialect.identifier_preparer
        quoted_fields = ", ".join(preparer.quote(name) for name in selected)
        quoted_table = (
            f"{preparer.quote(database)}." if database else ""
        ) + preparer.quote(table)
        sql = f"SELECT {quoted_fields} FROM {quoted_table}"
        run_sql, _, _ = engines.paginate_sql(sql, cfg.engine, limit + 1, 0)
        result = self._read(
            project,
            connection,
            cfg,
            run_sql,
            caller,
            limit,
            schema=database,
            mask=True,
        )
        data, media_type, ext = export_result(result["columns"], result["rows"], fmt)
        summary = {
            "project": project,
            "connection": connection,
            "database": database or cfg.database,
            "table": table,
            "fields": result["columns"],
            "row_count": result["row_count"],
            "requested_limit": limit,
            "truncated": result["truncated"],
            "format": fmt,
            "masked_columns": result.get("masked_columns", []),
        }
        return self._save_mcp_export(data, media_type, ext, summary)

    _MCP_EXPORT_TTL_S = 3600

    def _save_mcp_export(
        self, data: bytes, media_type: str, ext: str, summary: dict
    ) -> dict:
        """保存 MCP 导出产物并返回短期 bearer-token 下载链接。"""
        import re
        import secrets
        import json
        import time
        from datetime import UTC, datetime
        from pathlib import Path
        from urllib.parse import quote

        if not self.data_dir:
            raise QueryRejected("导出文件存储未启用（服务未配置 data_dir）")
        root = Path(self.data_dir) / "mcp_exports"
        root.mkdir(parents=True, exist_ok=True)
        now = time.time()
        self.purge_mcp_exports()

        token = secrets.token_hex(24)
        artifact_dir = root / token
        artifact_dir.mkdir(mode=0o700)
        raw_name = "_".join(
            str(part) for part in (
                summary.get("connection"), summary.get("database"), summary.get("table")
            ) if part
        ) or "quay_export"
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", raw_name).strip("._") or "quay_export"
        filename = f"{safe_name}.{ext}"
        path = artifact_dir / filename
        path.write_bytes(data)
        path.chmod(0o600)

        relative_url = f"/exports/{token}/{quote(filename)}"
        base = self.base_url.rstrip("/")
        result = {
            **summary,
            "token": token,
            "filename": filename,
            "media_type": media_type,
            "byte_size": len(data),
            "expires_at": int(now + self._MCP_EXPORT_TTL_S),
            "download_url": f"{base}{relative_url}" if base else relative_url,
            "agent_instruction": (
                "必要时用程序将 download_url 直接下载到目标位置；"
                "不要读取或把文件内容放入模型上下文。"
            ),
        }
        metadata = {
            **result,
            "created_at": datetime.fromtimestamp(now, UTC).isoformat(),
        }
        metadata_path = artifact_dir / "metadata.json"
        metadata_path.write_text(json.dumps(metadata, ensure_ascii=False), encoding="utf-8")
        metadata_path.chmod(0o600)
        return result

    def purge_mcp_exports(self) -> int:
        """删除超过 TTL 的临时导出目录，返回删除数量。"""
        import shutil
        import time
        from pathlib import Path

        if not self.data_dir:
            return 0
        root = Path(self.data_dir) / "mcp_exports"
        if not root.is_dir():
            return 0
        now = time.time()
        removed = 0
        for child in root.iterdir():
            try:
                if child.is_dir() and now - child.stat().st_mtime > self._MCP_EXPORT_TTL_S:
                    shutil.rmtree(child)
                    removed += 1
            except OSError:
                continue
        return removed

    def list_mcp_exports(self) -> list[dict]:
        """列出尚未过期的临时导出，供管理后台查看。"""
        import json
        import time
        from pathlib import Path

        self.purge_mcp_exports()
        if not self.data_dir:
            return []
        root = Path(self.data_dir) / "mcp_exports"
        if not root.is_dir():
            return []
        out = []
        for child in root.iterdir():
            try:
                meta = json.loads((child / "metadata.json").read_text(encoding="utf-8"))
                if int(meta.get("expires_at") or 0) <= int(time.time()):
                    continue
                # base_url 可能因重启/端口变化而改变，展示时用当前地址重新生成。
                filename = str(meta["filename"])
                relative = f"/exports/{child.name}/{filename}"
                meta["download_url"] = f"{self.base_url.rstrip('/')}{relative}" \
                    if self.base_url else relative
                out.append(meta)
            except (OSError, ValueError, KeyError, TypeError):
                continue
        return sorted(out, key=lambda item: item.get("created_at", ""), reverse=True)

    def delete_mcp_export(self, token: str) -> bool:
        """按随机 token 删除一条临时导出。"""
        import re
        import shutil
        from pathlib import Path

        if not self.data_dir or not re.fullmatch(r"[0-9a-f]{48}", token):
            return False
        path = Path(self.data_dir) / "mcp_exports" / token
        if not path.is_dir():
            return False
        shutil.rmtree(path)
        return True

    def preview_mcp_export(self, token: str, max_rows: int = 100) -> dict | None:
        """读取临时导出供人类后台预览；不会进入 MCP/agent 上下文。"""
        import csv
        import io
        import json
        import re
        from pathlib import Path

        if not self.data_dir or not re.fullmatch(r"[0-9a-f]{48}", token):
            return None
        artifact_dir = Path(self.data_dir) / "mcp_exports" / token
        try:
            meta = json.loads((artifact_dir / "metadata.json").read_text(encoding="utf-8"))
            filename = str(meta["filename"])
            path = self.resolve_mcp_export(token, filename)
            if path is None:
                return None
            fmt = str(meta.get("format") or "").lower()
            columns: list[str] = []
            rows: list[list] = []
            raw: str | None = None
            total = 0
            if fmt == "csv":
                parsed = list(csv.reader(io.StringIO(path.read_text(encoding="utf-8-sig"))))
                columns = parsed[0] if parsed else []
                total = max(0, len(parsed) - 1)
                rows = parsed[1:max_rows + 1]
            elif fmt == "json":
                data = json.loads(path.read_text(encoding="utf-8"))
                if isinstance(data, list) and data:
                    columns = list(data[0]) if isinstance(data[0], dict) else ["value"]
                    total = len(data)
                    rows = [
                        [item.get(c) for c in columns] if isinstance(item, dict) else [item]
                        for item in data[:max_rows]
                    ]
            elif fmt == "xlsx":
                from openpyxl import load_workbook

                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                iterator = ws.iter_rows(values_only=True)
                columns = [str(v or "") for v in next(iterator, ())]
                for row in iterator:
                    total += 1
                    if len(rows) < max_rows:
                        rows.append(list(row))
                wb.close()
            else:
                text = path.read_text(encoding="utf-8", errors="replace")
                raw = text[:50_000]
                total = len(text.splitlines())
            return {
                "metadata": meta,
                "columns": columns,
                "rows": rows,
                "raw": raw,
                "total_rows": total,
                "truncated": total > max_rows or (raw is not None and len(raw) < path.stat().st_size),
            }
        except (OSError, ValueError, KeyError, TypeError, json.JSONDecodeError):
            return None

    def resolve_mcp_export(self, token: str, filename: str):
        """校验短期下载 token，返回产物路径；任何异常都视为不存在。"""
        import re
        import time
        from pathlib import Path

        if not self.data_dir or not re.fullmatch(r"[0-9a-f]{48}", token):
            return None
        if filename != Path(filename).name:
            return None
        artifact_dir = Path(self.data_dir) / "mcp_exports" / token
        path = artifact_dir / filename
        try:
            if (
                not path.is_file()
                or time.time() - artifact_dir.stat().st_mtime > self._MCP_EXPORT_TTL_S
            ):
                return None
        except OSError:
            return None
        return path

    def admin_query_history(self, project: str, connection: str, limit: int = 30) -> list[dict]:
        """查询台历史面板：从审计取该连接最近执行过的 SQL，按文本去重保留最新。"""
        rows = self.store.recent(limit=300, filters={"project": project, "connection": connection})
        seen: set[str] = set()
        out: list[dict] = []
        for r in rows:
            sql = (r["sql"] or "").strip()
            if not sql or r["tool"] not in ("query", "execute", "admin_execute"):
                continue
            key = " ".join(sql.split()).lower()
            if key in seen:
                continue
            seen.add(key)
            out.append({"sql": sql, "ts": r["ts"], "status": r["status"], "tool": r["tool"]})
            if len(out) >= limit:
                break
        return out

    _EXPLAIN_PREFIX = {
        "mysql": "EXPLAIN FORMAT=JSON ",
        "postgres": "EXPLAIN (FORMAT JSON) ",
        "sqlite": "EXPLAIN QUERY PLAN ",
    }

    def admin_explain(
        self, project: str, connection: str, sql: str, caller: CallerInfo,
        schema: str | None = None,
    ) -> dict:
        """查询台 EXPLAIN：按引擎方言取执行计划（MySQL/PG 为 JSON，SQLite 为行）。

        纯 EXPLAIN 不执行语句（不带 ANALYZE），对写语句也安全。多语句拒绝。
        """
        import re as _re
        cfg = self.config.get_connection(project, connection)
        stmt = _re.sub(r"^\s*explain\s+", "", sql, flags=_re.IGNORECASE).strip().rstrip(";")
        if not stmt:
            raise QueryRejected("请先在编辑器写一条 SQL")
        verdict = classify(stmt, cfg.engine)
        if "多语句" in verdict.reason:
            raise QueryRejected("EXPLAIN 只支持单条语句")
        prefix = self._EXPLAIN_PREFIX.get(cfg.engine)
        if prefix is None:
            raise QueryRejected(f"引擎 {cfg.engine} 不支持 EXPLAIN")
        fmt = "rows" if cfg.engine == "sqlite" else "json"
        # 写语句（DELETE/UPDATE/INSERT/DDL）的 EXPLAIN 需要对应表的写权限：reader（只读账号）
        # 会被 DB 以权限不足拒绝（MySQL 1142）。EXPLAIN 不带 ANALYZE 不真正执行，改用 writer
        # 账号取计划是安全的；无独立 writer 时退回 reader（sqlite 等无账号概念场景）。
        role = "writer" if (not verdict.readonly and cfg.writer is not None) else "reader"

        def _run() -> dict:
            engine = self.pool.get(project, connection, cfg, role=role, schema=schema)
            # JSON 计划可能很长，放开单元格截断
            res = engines.run_query(engine, prefix + stmt, max_rows=500, max_cell_chars=1_000_000)
            return {"format": fmt, "columns": res.columns, "rows": res.rows}

        return self._audited(project, connection, cfg, "explain", stmt, caller, _run)

    # ---------- 分析工作台（DuckDB 沙箱，设计见 ANALYSIS.md）----------
    # 边界：工作区内任意 SQL 自由执行（本地草稿纸，不需审批）；
    # 从源库取数走 _read（reader 只读 + 审计 + 行数上限），生产红线不动。

    def _require_analysis(self):
        if self.analysis is None:
            raise QueryRejected("分析工作台未启用（需 serve 模式运行）")
        return self.analysis

    def _analysis_record(self, workspace: str, tool: str, sql: str, caller: CallerInfo) -> AuditRecord:
        return AuditRecord(project="analysis", connection=workspace, tool=tool, status="",
                           agent=caller.agent, session_id=caller.session_id,
                           environment="local", engine="duckdb", sql=sql)

    def analysis_overview(self) -> list[dict]:
        """工作区列表（含数据集摘要）。"""
        store = self._require_analysis()
        out = []
        for ws in store.list_workspaces():
            try:
                ws["datasets"] = store.list_datasets(ws["workspace"])
            except Exception:
                ws["datasets"] = []
            out.append(ws)
        return out

    def analysis_import(
        self, workspace: str, dataset: str, project: str, connection: str, source_sql: str,
        caller: CallerInfo, limit: int | None = None, schema: str | None = None,
    ) -> dict:
        """从某连接把查询结果快照进工作区（source_sql 也可为 `SELECT * FROM 表`）。

        只读校验 + 注入 LIMIT 上限 + reader 拉数（全程审计），随后落成 DuckDB 表。
        """
        from .analysis import DEFAULT_SNAPSHOT_ROWS, MAX_SNAPSHOT_ROWS
        store = self._require_analysis()
        cfg = self.config.get_connection(project, connection)
        if not classify(source_sql, cfg.engine).readonly:
            raise QueryRejected("快照导入仅支持只读查询（SELECT/SHOW/...）")
        n = min(limit or DEFAULT_SNAPSHOT_ROWS, MAX_SNAPSHOT_ROWS)
        run_sql, _, _ = engines.paginate_sql(source_sql, cfg.engine, n, 0)
        result = self._read(project, connection, cfg, run_sql, caller, n, schema=schema)
        spec = {"kind": "connection", "project": project, "connection": connection,
                "sql": source_sql, "limit": n, "schema": schema}
        imported = store.import_rows(workspace, dataset, result["columns"], result["rows"],
                                     spec=spec)
        rec = self._analysis_record(workspace, "analysis_import", source_sql, caller)
        rec.status = "ok"
        rec.detail = f"{project}/{connection} → {workspace}.{dataset}"
        rec.row_count = imported
        self.store.record(rec)
        return {"workspace": workspace, "dataset": dataset, "rows": imported,
                "truncated_to_limit": imported >= n}

    def analysis_import_file(
        self, workspace: str, dataset: str, path: str, caller: CallerInfo
    ) -> dict:
        store = self._require_analysis()
        rec = self._analysis_record(workspace, "analysis_import_file", path, caller)
        try:
            n = store.import_file(workspace, dataset, path)
        except Exception as e:
            rec.status = "error"
            rec.detail = f"{type(e).__name__}: {e}"
            self.store.record(rec)
            raise
        rec.status = "ok"
        rec.row_count = n
        self.store.record(rec)
        return {"workspace": workspace, "dataset": dataset, "rows": n}

    def analysis_sql(
        self, workspace: str, sql: str, caller: CallerInfo, max_rows: int | None = None
    ) -> dict:
        """在工作区执行任意 SQL（沙箱，自由写）。审计留痕。"""
        from .analysis import MAX_RESULT_ROWS
        store = self._require_analysis()
        rec = self._analysis_record(workspace, "analysis_sql", sql, caller)
        try:
            out = store.run_sql(workspace, sql, max_rows or MAX_RESULT_ROWS)
        except Exception as e:
            rec.status = "error"
            rec.detail = f"{type(e).__name__}: {e}"
            self.store.record(rec)
            raise
        rec.status = "ok"
        rec.row_count = out["row_count"]
        self.store.record(rec)
        return out

    # ---------- 分析 workflow（保存取数配方 + 脚本，一键重跑）----------

    def _require_workflows(self):
        if self.workflows is None:
            from .workflows import WorkflowError
            raise WorkflowError("workflow 存储未启用（需 serve 模式运行）")
        return self.workflows

    def workflow_save(self, name: str, workspace: str, script: str, caller: CallerInfo,
                      chart: dict | None = None, graph: dict | None = None,
                      allow_replace_graph: bool = True) -> dict:
        """保存 workflow：脚本/DAG + 取数配方 + 图表配置。

        DAG workflow 的取数配方在图的 source 节点里（编译时校验图合法）；
        纯脚本 workflow 从工作区 provenance 自动收集。
        allow_replace_graph=False（agent 侧）：同名 workflow 若是人画的 DAG，
        拒绝覆盖——agent 只允许创建/迭代脚本式 workflow。
        """
        from .workflows import compile_graph
        store = self._require_workflows()
        if not allow_replace_graph:
            existing = next((w for w in self.workflow_list() if w["name"] == name.strip()), None)
            if existing and existing.get("graph"):
                raise ValueError(
                    f"workflow {name!r} 是管理后台画布创建的 DAG，不允许覆盖；"
                    "请换一个名字，或让用户在后台修改")
        if graph:
            sources = compile_graph(graph)["sources"]  # 校验 + 配方以图为准
        else:
            sources = self._require_analysis().get_provenance(workspace)
        wf = store.save(name, workspace, script, sources, chart, graph)
        rec = self._analysis_record(workspace, "workflow_save", (script or "graph")[:500], caller)
        rec.status = "ok"
        rec.detail = f"workflow={name} sources={len(sources)} graph={bool(graph)}"
        self.store.record(rec)
        return wf.to_dict()

    def workflow_list(self) -> list[dict]:
        if self.workflows is None:
            return []
        return [w.to_dict() for w in self.workflows.list()]

    def workflow_delete(self, name: str) -> None:
        self._require_workflows().delete(name)

    def workflow_run(self, name: str, caller: CallerInfo) -> dict:
        """一键重跑：按 sources 重拉数据 → 逐条执行（脚本语句或 DAG 编译结果）→ 输出。

        任一步失败即停，标注在哪一步。全程审计（取数走 _read，脚本走 analysis_sql）。
        """
        from .workflows import compile_graph, split_statements
        wf = self._require_workflows().get(name)
        if wf.graph:
            plan = compile_graph(wf.graph)
            out = self._run_plan(wf.workspace, plan["sources"], plan["steps"], caller)
        else:
            stmts = [{"node": None, "name": f"步骤 {i}", "sql": s}
                     for i, s in enumerate(split_statements(wf.script), 1)]
            out = self._run_plan(wf.workspace, wf.sources, stmts, caller)
        return {"workflow": name, **out}

    def workflow_run_graph(self, workspace: str, graph: dict, caller: CallerInfo) -> dict:
        """直接运行画布上的 DAG（未保存也能跑）。编译失败作为第一步错误返回。"""
        from .workflows import WorkflowError, compile_graph
        try:
            plan = compile_graph(graph)
        except WorkflowError as e:
            return {"workflow": None, "ok": False, "output": None,
                    "steps": [{"step": "编译流程", "ok": False, "error": str(e)}]}
        return {"workflow": None, **self._run_plan(workspace, plan["sources"], plan["steps"], caller)}

    def workflow_preview_columns(self, workspace: str, graph: dict, node_id: str,
                                 caller: CallerInfo, refresh: bool = False) -> dict:
        """拿目标节点输出的列 schema（供上游 schema 感知用）。

        懒建策略：目标节点在工作区里已有对应 view/table 就直接 DESCRIBE；
        否则递归确保它依赖的 sources 已导入 + 前置 steps 已建 view，再 DESCRIBE。
        refresh=True 时强制重建（用于用户点「刷新 schema」）。
        编译异常或上游取数失败 → 返回 {columns:[], error:...} 而非抛（让前端友好提示）。
        """
        from .workflows import WorkflowError, compile_graph
        try:
            plan = compile_graph(graph)
        except WorkflowError as e:
            return {"columns": [], "error": f"编译流程失败：{e}"}
        # output 节点没有物化 view，改预览它的上游（"预览这一步输出" 等价于预览上游）
        target = _preview_target_for(graph, node_id)
        if target is None:
            return {"columns": [], "error": f"节点 {node_id!r} 不在流程中或未连接上游"}
        target_name, _target_node = target
        store = self._require_analysis()
        # 懒模式：目标已存在直接 DESCRIBE
        if not refresh and _dataset_exists(store, workspace, target_name):
            return {"columns": _describe_columns(store, workspace, target_name)}
        # 递归物化：只做目标依赖链上的 sources/steps
        needed = _plan_prefix_for(plan, target_name)
        for src in needed["sources"]:
            dataset = src["dataset"]
            if not refresh and _dataset_exists(store, workspace, dataset):
                continue
            try:
                if src.get("kind") == "file":
                    self.analysis_import_file(workspace, dataset, src["path"], caller)
                else:
                    self.analysis_import(workspace, dataset, src["project"], src["connection"],
                                         src["sql"], caller,
                                         limit=src.get("limit"), schema=src.get("schema"))
            except Exception as e:  # noqa: BLE001
                return {"columns": [], "error": f"上游节点「{dataset}」取数失败：{e}"}
        for st in needed["steps"]:
            step_name = st.get("name")
            if not step_name:
                continue
            if not refresh and _dataset_exists(store, workspace, step_name):
                continue
            try:
                self.analysis_sql(workspace, st["sql"], caller)
            except Exception as e:  # noqa: BLE001
                return {"columns": [], "error": f"节点「{step_name}」构建失败：{e}"}
        return {"columns": _describe_columns(store, workspace, target_name)}

    def workflow_preview_node(self, workspace: str, graph: dict, node_id: str,
                              caller: CallerInfo, limit: int = 100) -> dict:
        """预览目标节点的输出前 N 行（抽屉的「预览」标签用）。

        先确保依赖链已物化（复用 preview_columns 的懒模式），再 SELECT * LIMIT N。
        output 节点预览它的上游（output 不物化成 view）。
        返回 analysis_sql 的标准 dict（columns/rows/row_count）或 {error}。
        """
        cols_result = self.workflow_preview_columns(workspace, graph, node_id, caller)
        if cols_result.get("error"):
            return {"columns": [], "rows": [], "row_count": 0, "error": cols_result["error"]}
        target = _preview_target_for(graph, node_id)
        if target is None:
            return {"columns": [], "rows": [], "row_count": 0,
                    "error": f"节点 {node_id!r} 不在流程中或未连接上游"}
        target_name, _ = target
        try:
            n = max(1, min(int(limit or 100), 1000))
        except (TypeError, ValueError):
            n = 100
        return self.analysis_sql(workspace,
                                 f'SELECT * FROM "{target_name}" LIMIT {n}', caller)

    def _run_plan(self, workspace: str, sources: list[dict], steps: list[dict],
                  caller: CallerInfo) -> dict:
        """执行计划：重拉 sources → 顺序执行 steps（带 node id 供画布标注状态）。"""
        done: list[dict] = []
        for src in sources:
            label = f"导入 {src.get('dataset')}"
            node = src.get("node")
            try:
                if src.get("kind") == "file":
                    out = self.analysis_import_file(workspace, src["dataset"], src["path"], caller)
                else:
                    out = self.analysis_import(workspace, src["dataset"], src["project"],
                                               src["connection"], src["sql"], caller,
                                               limit=src.get("limit"), schema=src.get("schema"))
                done.append({"step": label, "node": node, "ok": True, "rows": out["rows"]})
            except Exception as e:  # noqa: BLE001
                done.append({"step": label, "node": node, "ok": False, "error": str(e)})
                return {"steps": done, "output": None, "ok": False}
        # output 保留旧语义（最后一个有 columns 的结果作主输出预览，向后兼容脚本式 workflow）
        # outputs 是新增的多 output 收集：只包含图上真正的 output 节点，供前端展示每个副产出
        output = None
        outputs: list[dict] = []
        for st in steps:
            label = f"{st['name']}: {st['sql'][:60]}"
            try:
                res = self.analysis_sql(workspace, st["sql"], caller)
                done.append({"step": label, "node": st.get("node"), "ok": True,
                             "rows": res["row_count"]})
                if res["columns"]:
                    output = res
                # 只把图里的 output 节点收集进 outputs（compile_graph 标了 is_output）
                if st.get("is_output") and res.get("columns"):
                    outputs.append({"node": st.get("node"), "name": st.get("name"), **res})
            except Exception as e:  # noqa: BLE001
                done.append({"step": label, "node": st.get("node"), "ok": False, "error": str(e)})
                return {"steps": done, "output": output, "outputs": outputs, "ok": False}
        return {"steps": done, "output": output, "outputs": outputs, "ok": True}

    # ---------- SQL 片段库（查询台保存/加载）----------

    def _require_snippets(self) -> "SnippetStore":
        if self.snippets is None:
            from .snippets import SnippetError
            raise SnippetError("片段库未启用")
        return self.snippets

    def list_snippets(self) -> list[dict]:
        if self.snippets is None:
            return []
        return [s.to_dict() for s in self.snippets.list()]

    def save_snippet(
        self, title: str, sql: str, note: str = "", connection: str = "",
        snippet_id: int | None = None,
    ) -> dict:
        store = self._require_snippets()
        if snippet_id is not None:
            return store.update(snippet_id, title, sql, note, connection).to_dict()
        return store.create(title, sql, note, connection).to_dict()

    def delete_snippet(self, snippet_id: int) -> None:
        self._require_snippets().delete(snippet_id)

    # ---------- 写操作（拒绝—重提 + change_id 放行）----------

    def execute(
        self,
        project: str,
        connection: str,
        sql: str,
        caller: CallerInfo,
        reason: str = "",
        change_id: int | None = None,
    ) -> dict:
        """写操作统一入口。

        - 只读语句：直接执行（等价于 query）；
        - 写操作 + 无 change_id：评估风险、生成审批单、拒绝并返回 change_id；
        - 写操作 + 有 change_id：校验审批单后执行**审批单里存储的 SQL**。
        """
        cfg = self.config.get_connection(project, connection)
        if self.approvals is None:
            raise QueryRejected("审批子系统未启用，无法执行写操作")

        # 带 change_id：一律走审批单核销（指纹校验 + 原子核销），不看重新分类结果——
        # 否则可构造「首提判写→生成审批单、重提判读→走 query() 绕开 consume 的指纹与核销」（H5）。
        if change_id is not None:
            return self._execute_approved(project, connection, cfg, sql, change_id, caller)

        verdict = classify(sql, cfg.engine)
        if verdict.readonly:
            return {"status": "executed", "readonly": True, **self.query(project, connection, sql, caller)}
        return self._request_approval(project, connection, cfg, sql, reason, caller)

    def _request_approval(
        self,
        project: str,
        connection: str,
        cfg: ConnectionConfig,
        sql: str,
        reason: str,
        caller: CallerInfo,
    ) -> dict:
        report = assess(sql, cfg.engine, self._meta_provider(project, connection, cfg))
        report_dict = report.to_dict()
        plan = self._try_explain(project, connection, cfg, sql)
        if plan:
            report_dict["explain"] = plan
        change = self.approvals.create(
            project=project,
            connection=connection,
            environment=cfg.environment,
            engine=cfg.engine,
            sql=sql,
            fingerprint=fingerprint(sql, cfg.engine),
            reason=reason,
            risk_level=report.level,
            risk_report=report_dict,
            agent=caller.agent,
            session_id=caller.session_id,
        )
        rec = self._base_record(project, connection, cfg, "execute", sql, caller)
        rec.status = "rejected"
        rec.detail = f"需人工授权，已生成审批单 #{change.id}（风险 {report.level}）"
        self.store.record(rec)
        # 需要人为介入 → 主动发通知（安静即正常：不通知的话可能长时间没人看到）
        # meta.deeplink 让各渠道适配跳转：Bark→url 字段、企微→markdown 链接、
        # 飞书→post 富文本 a 节点、macOS→body 附 URL 文本、站内 inbox→前端点击
        try:
            from .notify import approval_deeplink  # noqa: PLC0415
            sql_preview = " ".join(sql.split())[:120]
            base_url = str(self._setting("admin_base_url") or "http://127.0.0.1:8100")
            self.notifier.send(
                title=f"新审批单 #{change.id} · {project}/{connection}",
                body=f"风险 {report.level} · agent={caller.agent or 'unknown'}\nSQL: {sql_preview}",
                meta={"kind": "approval_created", "change_id": change.id,
                      "project": project, "connection": connection,
                      "risk_level": report.level,
                      "deeplink": approval_deeplink(base_url, change.id)},
            )
        except Exception:  # noqa: BLE001
            logger.exception("notify approval_created failed")
        return {
            "status": "approval_required",
            "change_id": change.id,
            "risk": report_dict,
            "message": (
                f"该操作被评估为需人工授权（风险等级 {report.level}）。"
                f"已生成审批单 #{change.id}，请通知用户在管理后台审批；"
                f"批准后带上 change_id={change.id} 重新提交相同 SQL 即可执行。"
                f"审批单 60 分钟内有效。"
            ),
        }

    def _execute_approved(
        self,
        project: str,
        connection: str,
        cfg: ConnectionConfig,
        sql: str,
        change_id: int,
        caller: CallerInfo,
    ) -> dict:
        rec = self._base_record(project, connection, cfg, "execute", sql, caller)
        try:
            change = self.approvals.consume(
                change_id, fingerprint(sql, cfg.engine), (project, connection)
            )
        except ApprovalError as e:
            rec.status = "rejected"
            rec.detail = str(e)
            self.store.record(rec)
            return {"status": "rejected", "change_id": change_id, "reason": str(e)}

        # 执行审批单里存储的 SQL（不是 agent 重提的文本），用 writer 账号

        def _do() -> "engines.QueryResult":
            engine = self.pool.get(project, connection, cfg, role="writer")
            return engines.run_write(engine, change.sql)

        try:
            result = self._run_touching_db(project, connection, _do)
        except ConnectionUnavailable as e:
            rec.status = "error"
            rec.detail = f"ConnectionUnavailable[{e.state}]: {e}"
            self.store.record(rec)
            raise
        except Exception as e:
            rec.status = "error"
            rec.detail = f"{type(e).__name__}: {e}"
            self.store.record(rec)
            raise

        rec.status = "ok"
        rec.detail = f"审批单 #{change_id} 已核销（审批人 {change.decided_by}）"
        rec.row_count = result.row_count
        rec.duration_ms = result.duration_ms
        self.store.record(rec)
        return {
            "status": "executed",
            "change_id": change_id,
            "affected_rows": result.row_count,
            "duration_ms": result.duration_ms,
        }

    def _try_explain(
        self, project: str, connection: str, cfg: ConnectionConfig, sql: str,
        schema: str | None = None,
    ) -> str | None:
        """对写语句取执行计划（不带 ANALYZE，不执行）供审批人参考。

        reader 会话可能因只读事务拒绝 EXPLAIN DML（PG 会），失败则退回 writer；
        全部失败返回 None，不阻断审批单生成。计划文本截断到 4000 字符。
        """
        for role in ("reader", "writer"):
            if role == "writer" and cfg.writer is None:
                break
            try:
                engine = self.pool.get(project, connection, cfg, role=role, schema=schema)
            except Exception:
                continue
            plan = engines.explain(engine, sql, cfg.engine)
            if plan:
                return plan[:4000]
        return None

    def _meta_provider(self, project: str, connection: str, cfg: ConnectionConfig):
        """给风险引擎注入"按表取元数据"的能力；无缓存或取不到时返回 None。"""
        if self.metadata is None:
            return lambda _table: None

        def provider(table: str):
            try:
                return self.metadata.get(project, connection, cfg, table)
            except Exception:
                return None

        return provider

    # ---------- 系统设置（后台界面偏好）----------

    def get_settings(self) -> dict:
        from .settings import DEFAULTS
        return self.settings.get_all() if self.settings is not None else dict(DEFAULTS)

    def save_settings(self, updates: dict) -> dict:
        if self.settings is None:
            raise QueryRejected("设置子系统未启用")
        return self.settings.save(updates)

    def _setting(self, key: str):
        return self.get_settings().get(key)

    def agent_result_budget(self, project: str, connection: str) -> int:
        """解析给 agent 的结果字符预算：连接级 Policy 优先，否则全局设置兜底。"""
        cfg = self.config.get_connection(project, connection)
        if cfg.policy.agent_max_result_chars:
            return int(cfg.policy.agent_max_result_chars)
        return int(self._setting("agent_max_result_chars") or DEFAULT_AGENT_MAX_RESULT_CHARS)

    # ---------- Redis 浏览 / 命令窗口（管理后台，对标 Medis）----------

    def _redis_cfg(self, project: str, connection: str) -> ConnectionConfig:
        cfg = self.config.get_connection(project, connection)
        if cfg.engine != "redis":
            raise QueryRejected(f"连接 {project}/{connection} 引擎为 {cfg.engine}，不是 Redis")
        return cfg

    def redis_databases(self, project: str, connection: str, caller: CallerInfo) -> list[dict]:
        """列出全部逻辑库（db0..N-1），有数据的带键数。对标 Medis 底部库切换器。"""
        cfg = self._redis_cfg(project, connection)
        min_dbs = int(self._setting("redis_min_dbs") or redis_engine.MIN_DBS_SHOWN)
        return self._audited(
            project, connection, cfg, "redis_keyspace", "", caller,
            lambda: redis_engine.keyspace_dbs(self.redis_pool.get(project, connection, cfg),
                                              min_dbs=min_dbs))

    def redis_keys(
        self, project: str, connection: str, caller: CallerInfo,
        db: int | None = None, pattern: str = "*", max_keys: int | None = None,
    ) -> dict:
        cfg = self._redis_cfg(project, connection)
        limit = max_keys if max_keys is not None else int(self._setting("redis_key_limit"))
        scan_count = int(self._setting("redis_scan_count") or 500)
        detail = f"db={db if db is not None else ''} match={pattern}"
        return self._audited(
            project, connection, cfg, "redis_scan", detail, caller,
            lambda: redis_engine.scan_keys(
                self.redis_pool.get(project, connection, cfg, db=db),
                pattern=pattern or "*", max_keys=limit, scan_count=scan_count))

    def redis_value(
        self, project: str, connection: str, key: str, caller: CallerInfo,
        db: int | None = None,
    ) -> dict:
        cfg = self._redis_cfg(project, connection)
        redis_engine.set_msgpack_decode(bool(self._setting("redis_msgpack_decode")))
        detail = f"db={db if db is not None else ''} key={key}"
        return self._audited(
            project, connection, cfg, "redis_read", detail, caller,
            lambda: redis_engine.read_value(
                self.redis_pool.get(project, connection, cfg, db=db), key,
                max_cell_chars=cfg.policy.max_cell_chars))

    def admin_redis_run(
        self, project: str, connection: str, command: str, caller: CallerInfo,
        confirm: bool = False, db: int | None = None, confirm_text: str | None = None,
    ) -> dict:
        """后台命令窗口专用入口（对标 admin_run_sql 的 Redis 版）。

        - 读命令：直通 reader 出结果。
        - 写命令 + confirm=False：返回风险报告，不执行。
        - 写命令 + confirm=True：writer（无 writer 则 reader）直接执行并审计（tool=admin_execute）。
          Redis 只供人通过后台操作（不暴露给 agent），故无 agent 侧审批流。
        - **生产环境写命令**：单次确认不够，须额外输入连接名（confirm_text）匹配才放行，
          防误清线上库（对齐 SQL 侧 prod 强管控）。
        """
        cfg = self._redis_cfg(project, connection)
        is_prod = (cfg.environment or "").lower() == "prod"
        verdict = classify_command(command)
        parts = parse_command(command)
        # 命令原文脱敏后再入审计（密码永不进审计记录）；执行仍用未脱敏的 parts
        safe_command = redis_engine.redact_command_text(command, parts)

        if verdict.readonly:
            rec = self._base_record(project, connection, cfg, "redis_command", safe_command, caller)
            rec.fingerprint = command_fingerprint(safe_command)
            if db is not None:
                rec.detail = f"db={db}"

            def _do_read():
                client = self.redis_pool.get(project, connection, cfg, db=db)
                return redis_engine.run_command(client, parts,
                                                max_cell_chars=cfg.policy.max_cell_chars)

            try:
                result = self._run_touching_db(project, connection, _do_read)
            except ConnectionUnavailable as e:
                rec.status = "error"
                rec.detail = f"ConnectionUnavailable[{e.state}]: {e}"
                self.store.record(rec)
                raise
            except Exception as e:
                rec.status = "error"
                rec.detail = f"{type(e).__name__}: {e}"
                self.store.record(rec)
                raise
            rec.status = "ok"
            rec.duration_ms = result.duration_ms
            self.store.record(rec)
            return {"kind": "read", "readonly": True, "command": verdict.command,
                    "value": result.value, "duration_ms": result.duration_ms}

        if not confirm:
            return {"kind": "confirm", "statement_kind": f"Redis:{verdict.command}",
                    "prod": is_prod, "expect_text": connection if is_prod else None,
                    "risk": {"level": verdict.level, "statement_kind": f"Redis:{verdict.command}",
                             "tables": [], "reasons": [verdict.reason], "warnings": []}}

        # 生产环境：确认之外还须输入连接名匹配，否则拒绝执行
        if is_prod and (confirm_text or "").strip() != connection:
            raise QueryRejected(
                f"生产环境写命令需输入连接名「{connection}」确认后才执行")

        rec = self._base_record(project, connection, cfg, "admin_execute", safe_command, caller)
        rec.fingerprint = command_fingerprint(safe_command)
        role = "writer" if cfg.writer is not None else "reader"

        def _do_write():
            client = self.redis_pool.get(project, connection, cfg, role=role, db=db)
            return redis_engine.run_command(client, parts,
                                            max_cell_chars=cfg.policy.max_cell_chars)

        try:
            result = self._run_touching_db(project, connection, _do_write)
        except ConnectionUnavailable as e:
            rec.status = "error"
            rec.detail = f"ConnectionUnavailable[{e.state}]: {e}"
            self.store.record(rec)
            raise
        except Exception as e:
            rec.status = "error"
            rec.detail = f"{type(e).__name__}: {e}"
            self.store.record(rec)
            raise
        rec.status = "ok"
        rec.detail = "后台命令窗口直接执行（已二次确认）" + (f" db={db}" if db is not None else "")
        rec.duration_ms = result.duration_ms
        self.store.record(rec)
        return {"kind": "write", "command": verdict.command,
                "value": result.value, "duration_ms": result.duration_ms}

    # ---------- 审批决策（管理后台 / elicitation 调用）----------

    def approve_change(self, change_id: int, decided_by: str, note: str = ""):
        if self.approvals is None:
            raise QueryRejected("审批子系统未启用")
        return self.approvals.approve(change_id, decided_by, note)

    def reject_change(self, change_id: int, decided_by: str, note: str = ""):
        if self.approvals is None:
            raise QueryRejected("审批子系统未启用")
        return self.approvals.reject(change_id, decided_by, note)

    def get_change(self, change_id: int):
        if self.approvals is None:
            raise QueryRejected("审批子系统未启用")
        return self.approvals.get(change_id)

    def list_changes(self, status: str | None = None):
        if self.approvals is None:
            return []
        return self.approvals.list_by_status(status)

    # ---------- schema 探索 ----------

    def list_databases(self, project: str, connection: str, caller: CallerInfo) -> list[str]:
        """列出连接可选的库/schema（MySQL 数据库 / PG schema）。sqlite 无此概念返回 []。"""
        cfg = self.config.get_connection(project, connection)
        if cfg.engine not in ("mysql", "postgres", "clickhouse"):
            return []
        engine = self.pool.get(project, connection, cfg)
        return self._audited(project, connection, cfg, "list_databases", "", caller,
                             lambda: engines.list_databases(engine))

    def list_tables(
        self, project: str, connection: str, caller: CallerInfo, schema: str | None = None
    ) -> list[str]:
        cfg = self.config.get_connection(project, connection)
        # 未绑定默认库时，先让用户选库（库→表→列 三级树）。MySQL/PG 不带 schema 反射会崩
        # （默认 schema 为 None）；ClickHouse 不会崩但会落到 default 库、看不到别的库 → 一并引导
        if schema is None and not cfg.database and cfg.engine in ("mysql", "postgres", "clickhouse"):
            raise ValueError("此连接未绑定默认库，请先选择一个库（schema）再列表")
        engine = self.pool.get(project, connection, cfg)
        return self._audited(project, connection, cfg, "list_tables", schema or "", caller,
                             lambda: engines.list_tables(engine, schema))

    def describe_table(
        self, project: str, connection: str, table: str, caller: CallerInfo,
        schema: str | None = None,
    ) -> dict:
        cfg = self.config.get_connection(project, connection)
        # 未绑定默认库时的两道防线（否则 SQLAlchemy 反射取默认库为 None → NoneType.replace 崩）：
        # ① 从「库.表」限定名拆出 schema；② 仍无 schema 且无默认库 → 明确报错引导，而非让它崩
        if schema is None and "." in table:
            schema, table = table.split(".", 1)
        if schema is None and not cfg.database and cfg.engine in ("mysql", "postgres", "clickhouse"):
            raise ValueError("此连接未绑定默认库，请用「库名.表名」指定表，或先选择一个库（schema）")
        engine = self.pool.get(project, connection, cfg)
        detail = f"{schema}.{table}" if schema else table
        return self._audited(project, connection, cfg, "describe_table", detail, caller,
                             lambda: engines.describe_table(engine, table, schema))

    def admin_search_tables(
        self, project: str, connection: str, q: str, caller: CallerInfo,
    ) -> list[dict]:
        """查询台全局表名搜索（⌘P）：跨库 LIKE 匹配，最多 50 条。"""
        q = (q or "").strip()
        if not q:
            return []
        cfg = self.config.get_connection(project, connection)
        engine = self.pool.get(project, connection, cfg)
        return self._audited(project, connection, cfg, "search_tables", q, caller,
                             lambda: engines.search_tables(engine, cfg.engine, q))

    def admin_table_sizes(
        self, project: str, connection: str, caller: CallerInfo, schema: str | None = None
    ) -> dict[str, int]:
        """查询台树右侧的表容量（字节）。取不到返回空 dict，不阻断列表。"""
        cfg = self.config.get_connection(project, connection)
        engine = self.pool.get(project, connection, cfg)
        return self._audited(project, connection, cfg, "table_sizes", schema or "", caller,
                             lambda: engines.table_sizes(engine, cfg.engine, schema))

    def get_table_ddl(
        self, project: str, connection: str, table: str, caller: CallerInfo,
        schema: str | None = None,
    ) -> str:
        """取建表语句（查询台「查看 DDL」）。"""
        cfg = self.config.get_connection(project, connection)
        engine = self.pool.get(project, connection, cfg)
        detail = f"{schema}.{table}" if schema else table
        return self._audited(project, connection, cfg, "table_ddl", detail, caller,
                             lambda: engines.get_table_ddl(engine, cfg.engine, table, schema))

    def ai_generate_sql(
        self, project: str, connection: str, question: str, caller: CallerInfo,
        *, schema: str | None = None, tables: list[str] | None = None,
        explain: bool = False, include_samples: bool = False,
        session_id: str | None = None,
    ) -> dict:
        """让命令行 AI 按表结构 + 自然语言需求生成一条 SQL。只生成、不执行。

        tables 为空 = 「整库」模式：列出该库的表（超 ai_max_tables 报错要求收窄）。
        include_samples 时附少量样本行帮助 AI 理解数据形态。
        session_id 非空 = 追问：续接同一会话、不重发表结构。返回 {sql, explanation, session_id}。
        """
        from . import ai

        s = self.get_settings()
        if not s.get("ai_enabled"):
            raise QueryRejected("AI 辅助未开启，请在系统设置中开启")
        question = (question or "").strip()
        if not question:
            raise QueryRejected("请填写你想查什么")
        cfg = self.config.get_connection(project, connection)
        if cfg.engine not in ("mysql", "postgres", "sqlite"):
            raise QueryRejected(f"连接引擎 {cfg.engine} 暂不支持 AI 生成 SQL")
        engine = self.pool.get(project, connection, cfg)
        max_tables = int(s.get("ai_max_tables") or 40)

        def _run() -> dict:
            ddls: list[tuple[str, str]] = []
            samples: dict[str, str] | None = None
            if not session_id:  # 首轮才收集表结构；追问续接会话、上下文已在 AI 侧
                names = list(tables or [])
                if not names:  # 整库：列出全部表
                    names = engines.list_tables(engine, schema)
                if not names:
                    raise QueryRejected("该库没有可用的表")
                if len(names) > max_tables:
                    raise QueryRejected(
                        f"待发送的表有 {len(names)} 张，超过上限 {max_tables}；请勾选具体的表，"
                        "或在系统设置调大「最大表数」")
                for t in names:
                    tbl_schema, tbl = (t.split(".", 1) if "." in t else (schema, t))
                    ddls.append((tbl, engines.get_table_ddl(engine, cfg.engine, tbl, tbl_schema)))
                if include_samples:
                    samples = {}
                    for t in names:
                        tbl_schema, tbl = (t.split(".", 1) if "." in t else (schema, t))
                        try:
                            r = engines.sample_rows(engine, tbl, 5,
                                                    max_cell_chars=cfg.policy.max_cell_chars)
                            samples[tbl] = _rows_to_text(r.columns, r.rows)
                        except Exception:  # 样本拿不到不阻断生成
                            continue
            result = ai.generate_sql(
                system_prompt=str(s.get("ai_sql_prompt") or ai.DEFAULT_SQL_PROMPT),
                dialect=cfg.engine, ddls=ddls, question=question,
                explain=explain, samples=samples,
                provider=str(s.get("ai_provider") or "claude"),
                model=str(s.get("ai_model") or ""),
                timeout=int(s.get("ai_timeout_s") or 60),
                cli_path=str(s.get("ai_cli_path") or ""),
                session_id=session_id, api=_ai_api_cfg(s))
            if not result.sql.strip():
                raise QueryRejected("AI 未能生成 SQL，请补充需求描述后重试")
            return {"sql": result.sql, "explanation": result.explanation,
                    "session_id": result.session_id}

        tool = "ai_followup_sql" if session_id else "ai_generate_sql"
        try:
            return self._audited(project, connection, cfg, tool,
                                 question[:2000], caller, _run)
        except ai.AIError as e:
            raise QueryRejected(str(e)) from e

    def ai_generate_workflow(
        self, project: str, connection: str, question: str, caller: CallerInfo,
        *, schema: str | None = None, tables: list[str] | None = None,
    ) -> dict:
        """让命令行 AI 按连接/表结构 + 需求设计一张 workflow DAG（画布图）。

        产物用 compile_graph 校验，编译失败把错误回喂给 AI 重修一次；仍失败则报错。
        返回 {graph:{nodes,edges}}（节点已排版赋 x/y），前端载到画布待人审阅、不自动执行。
        """
        from . import ai
        from .workflows import WorkflowError, compile_graph

        s = self.get_settings()
        if not s.get("ai_enabled"):
            raise QueryRejected("AI 辅助未开启，请在系统设置中开启")
        question = (question or "").strip()
        if not question:
            raise QueryRejected("请描述你想做的分析流程")
        cfg = self.config.get_connection(project, connection)
        if cfg.engine not in ("mysql", "postgres", "sqlite"):
            raise QueryRejected(f"连接引擎 {cfg.engine} 暂不支持 AI 生成流程")
        engine = self.pool.get(project, connection, cfg)
        max_tables = int(s.get("ai_max_tables") or 40)
        # 可用连接（供 source 节点选，排除 redis）
        conns = [f"{p}/{c}" for p, proj in sorted(self.config.projects.items())
                 for c, cc in sorted(proj.connections.items()) if cc.engine != "redis"]

        def _run() -> dict:
            names = list(tables or [])
            if not names:
                names = engines.list_tables(engine, schema)
            if len(names) > max_tables:
                raise QueryRejected(
                    f"待发送的表有 {len(names)} 张，超过上限 {max_tables}；请勾选具体的表")
            ddls: list[tuple[str, str]] = []
            for t in names:
                tbl_schema, tbl = (t.split(".", 1) if "." in t else (schema, t))
                ddls.append((tbl, engines.get_table_ddl(engine, cfg.engine, tbl, tbl_schema)))
            kw = dict(system_prompt=str(s.get("ai_workflow_prompt") or ai.DEFAULT_WORKFLOW_PROMPT),
                      dialect=cfg.engine, connections=conns, ddls=ddls, question=question,
                      provider=str(s.get("ai_provider") or "claude"),
                      model=str(s.get("ai_model") or ""),
                      timeout=int(s.get("ai_timeout_s") or 60),
                      cli_path=str(s.get("ai_cli_path") or ""), api=_ai_api_cfg(s))
            graph, sid = ai.generate_workflow(**kw)
            try:
                compile_graph(graph)
            except WorkflowError as e:  # 回喂错误、续接会话重修一次
                graph, sid = ai.generate_workflow(**kw, repair_error=str(e), session_id=sid)
                try:
                    compile_graph(graph)
                except WorkflowError as e2:
                    raise QueryRejected(f"AI 生成的流程仍不合法：{e2}") from e2
            _layout_graph(graph)
            return {"graph": graph}

        try:
            return self._audited(project, connection, cfg, "ai_generate_workflow",
                                 question[:2000], caller, _run)
        except ai.AIError as e:
            raise QueryRejected(str(e)) from e

    def sample_rows(self, project: str, connection: str, table: str, limit: int, caller: CallerInfo) -> dict:
        cfg = self.config.get_connection(project, connection)
        limit = min(limit, cfg.policy.max_rows)
        engine = self.pool.get(project, connection, cfg)

        def _run() -> dict:
            result = engines.sample_rows(engine, table, limit,
                                         max_cell_chars=cfg.policy.max_cell_chars)
            rows, masked = apply_mask(result.columns, result.rows, cfg.policy)
            out = {
                "columns": result.columns,
                "rows": rows,
                "row_count": result.row_count,
                "duration_ms": result.duration_ms,
            }
            if masked:
                out["masked_columns"] = masked
            return out

        return self._audited(project, connection, cfg, "sample_rows", table, caller, _run)

    def test_connection(self, project: str, connection: str, caller: CallerInfo) -> dict:
        cfg = self.config.get_connection(project, connection)

        def _run() -> dict:
            engine = self.pool.get(project, connection, cfg)
            result = engines.run_query(engine, "SELECT 1", max_rows=1)
            return {"ok": True, "engine": cfg.engine, "duration_ms": result.duration_ms}

        return self._audited(project, connection, cfg, "test_connection", "", caller, _run)

    def reconnect_connection(self, project: str, connection: str, caller: CallerInfo) -> dict:
        """人工触发的强制重连：清健康位 + 回收旧引擎/隧道，然后主动探测重建。

        与 test_connection 的关键区别：test_connection 走 `_run_touching_db`，连接处于
        unavailable/exhausted 时会被 `health.check()` 直接挡下、根本不会尝试重建；exhausted
        连接因此无法靠它自愈。这里**绕过健康位**，无条件重置后走 `_health_probe`（SELECT 1 /
        PING）建立新连接——供管理后台/查询台在收到「连接不可用」错误时点「重连」使用。

        探测成功：健康位已随成功清为 ok，返回 {ok, engine}。
        探测失败：记审计 + 若为连接级错误则 mark_failed（后台退避重连从头再来），原样抛出。
        """
        cfg = self.config.get_connection(project, connection)  # 连接不存在 → KeyError
        # 无条件清健康状态：exhausted/unavailable 都归零，让本次探测不被拦
        self.health.force_clear(project, connection)
        rec = self._base_record(project, connection, cfg, "reconnect", "", caller)
        try:
            # _health_probe 会先 dispose 旧引擎/隧道再探测，拿到的是全新连接
            self._health_probe(project, connection)
        except Exception as e:  # noqa: BLE001
            rec.status = "error"
            rec.detail = f"{type(e).__name__}: {e}"
            self.store.record(rec)
            if is_connection_error(e):
                # 仍连不上：重新进入后台退避重连（给它再一轮机会）
                self.health.mark_failed(project, connection, f"{type(e).__name__}: {e}")
            raise
        rec.status = "ok"
        rec.detail = "重连成功"
        self.store.record(rec)
        return {"ok": True, "engine": cfg.engine}

    # ---------- 内部 ----------

    def _base_record(
        self,
        project: str,
        connection: str,
        cfg: ConnectionConfig,
        tool: str,
        sql: str,
        caller: CallerInfo,
    ) -> AuditRecord:
        return AuditRecord(
            project=project,
            connection=connection,
            tool=tool,
            status="",
            agent=caller.agent,
            session_id=caller.session_id,
            environment=cfg.environment,
            engine=cfg.engine,
            sql=sql,
            fingerprint=fingerprint(sql, cfg.engine) if sql else "",
        )

    def _audited(self, project, connection, cfg, tool, detail_sql, caller, fn):  # noqa: ANN001
        rec = self._base_record(project, connection, cfg, tool, detail_sql, caller)
        try:
            result = self._run_touching_db(project, connection, fn)
        except ConnectionUnavailable as e:
            rec.status = "error"
            rec.detail = f"ConnectionUnavailable[{e.state}]: {e}"
            self.store.record(rec)
            raise
        except Exception as e:
            rec.status = "error"
            rec.detail = f"{type(e).__name__}: {e}"
            self.store.record(rec)
            raise
        rec.status = "ok"
        self.store.record(rec)
        return result

    def _run_touching_db(self, project: str, connection: str, fn):  # noqa: ANN001
        """任何"会触达 DB/隧道"的动作都过这里：入口先查健康位、出错时按类别打标。

        - 若健康位为 unavailable/exhausted：直接抛 ConnectionUnavailable（不碰 DB）
        - 执行成功：清健康标记（如果之前挂过）
        - 失败：判断是不是"连接级"异常，是就打标 + 启后台重连，然后原样再抛
          （非连接级异常如 SQL 语法/权限拒/审批拒不打标，重连也没用）
        """
        self.health.check(project, connection)
        try:
            result = fn()
        except ConnectionUnavailable:
            raise
        except Exception as e:
            if is_connection_error(e):
                # 池里对应的引擎/隧道大概率也坏了：回收让重连时用新连接
                try:
                    self.pool.dispose_connection(project, connection)
                    self.redis_pool.dispose_connection(project, connection)
                except Exception:  # noqa: BLE001
                    pass
                self.health.mark_failed(project, connection, f"{type(e).__name__}: {e}")
            raise
        self.health.mark_ok(project, connection)
        return result

    def _health_probe(self, project: str, connection: str) -> None:
        """健康监控的探测回调：走 reader 建/借连接做 SELECT 1（Redis 用 PING）。

        失败原样抛出，让 HealthMonitor 记退避；成功即视作连接已恢复。
        """
        try:
            cfg = self.config.get_connection(project, connection)
        except KeyError:
            # 连接被删了：视作已恢复（后续不会再有请求走它）
            return
        # 每次探测前先回收旧引擎/隧道，避免复用坏连接
        try:
            self.pool.dispose_connection(project, connection)
            self.redis_pool.dispose_connection(project, connection)
        except Exception:  # noqa: BLE001
            pass
        if cfg.engine == "redis":
            client = self.redis_pool.get(project, connection, cfg)
            client.ping()
            return
        engine = self.pool.get(project, connection, cfg)
        engines.run_query(engine, "SELECT 1", max_rows=1)

    def _on_connection_exhausted(self, project: str, connection: str, error: str) -> None:
        """连接 exhausted 事件回调：只落 warn 日志，不发通知。

        原因：连接不可用会在 agent 侧被 `[connection_exhausted]` ToolError 直接告知，
        agent 会告诉用户；再发桌面/群通知反而形成噪音（尤其自建 server 抖动时会连发）。
        审批单等"必须人主动介入"的场景仍走通知（那里 agent 不再触达）。
        """
        logger.warning("connection %s/%s exhausted: %s", project, connection, error)

    # ---------- 连接管理（管理后台，需已配置 config_path）----------

    def _require_config_path(self) -> str:
        if not self.config_path:
            raise QueryRejected("未设置配置文件路径，无法在线管理连接")
        return self.config_path

    def upsert_connection(self, project: str, connection: str, caller: CallerInfo, **fields) -> None:
        from .connections import ConnectionManager

        mgr = ConnectionManager(self.config, self._require_config_path())
        mgr.upsert(project, connection, **fields)
        self._after_connection_change(project, connection, caller, "upsert_connection",
                                      f"引擎 {fields.get('engine')}")

    # ---------- SSH 证书库 ----------

    def list_ssh_identities(self) -> dict:
        return dict(self.config.ssh_identities)

    def upsert_ssh_identity(
        self, name: str, key_path: str, known_hosts_path: str | None, caller: CallerInfo,
        host: str | None = None, user: str | None = None, port: str | int | None = None,
    ) -> None:
        from .connections import ConnectionManager

        mgr = ConnectionManager(self.config, self._require_config_path())
        referers = mgr.identity_referers(name)
        mgr.upsert_identity(name, key_path, known_hosts_path, host=host, user=user, port=port)
        # 证书变更影响引用它的连接的隧道：回收让下次用新证书重建
        for ref in referers:
            proj, conn = ref.split("/", 1)
            self.pool.dispose_connection(proj, conn)
            self.redis_pool.dispose_connection(proj, conn)
        self.store.record(AuditRecord(
            project="admin", connection=name, tool="upsert_ssh_identity", status="ok",
            agent=caller.agent, session_id=caller.session_id, detail="已保存 SSH 配置"))

    def delete_ssh_identity(self, name: str, caller: CallerInfo) -> None:
        from .connections import ConnectionManager

        mgr = ConnectionManager(self.config, self._require_config_path())
        mgr.delete_identity(name)
        self.store.record(AuditRecord(
            project="admin", connection=name, tool="delete_ssh_identity", status="ok",
            agent=caller.agent, session_id=caller.session_id, detail="已删除 SSH 配置"))

    def probe_connection_fields(self, fields: dict, existing_password: str | None = None):
        """用表单值临时探测连通性与账号权限（测试按钮）。不保存、不入池。"""
        from .config import ConnectionConfig, Policy
        from .probe import probe_connection

        password = fields.get("password") or None
        eff_pw = f"plain://{password}" if password else existing_password
        # sqlite 无账号；redis 允许无认证（本地无 auth 实例）——都不强制填密码
        if eff_pw is None and fields.get("engine") not in ("sqlite", "redis"):
            from .probe import ProbeResult
            return ProbeResult(ok=False, message="请填写密码后再测试")
        cfg = ConnectionConfig(
            engine=fields["engine"], environment=fields.get("environment", "dev"),
            host=fields.get("host") or None, port=fields.get("port"),
            database=fields.get("database") or None, user=fields.get("user") or None,
            password=eff_pw, jump_hosts=fields.get("jump_hosts", []),
            ssh_options=fields.get("ssh_options", []),
            policy=Policy(max_rows=fields.get("max_rows", 500)),
        )
        return probe_connection(cfg, None, self.config.ssh_identities)

    def probe_ssh_fields(self, fields: dict):
        """用表单值只测 SSH 跳板链是否可建隧道。"""
        from .config import ConnectionConfig
        from .probe import probe_ssh

        # SSH 测试只用 host/port/跳板，user/password 填占位满足校验
        cfg = ConnectionConfig(
            engine=fields["engine"], environment=fields.get("environment", "dev"),
            host=fields.get("host") or "127.0.0.1", port=fields.get("port"),
            database=fields.get("database") or None, user="_probe",
            password="plain://_", jump_hosts=fields.get("jump_hosts", []),
            ssh_options=fields.get("ssh_options", []),
        )
        return probe_ssh(cfg, self.config.ssh_identities)

    def delete_connection(self, project: str, connection: str, caller: CallerInfo) -> None:
        from .connections import ConnectionManager

        mgr = ConnectionManager(self.config, self._require_config_path())
        mgr.delete(project, connection)
        self._after_connection_change(project, connection, caller, "delete_connection", "已删除")

    def _after_connection_change(
        self, project: str, connection: str, caller: CallerInfo, tool: str, detail: str
    ) -> None:
        # 回收旧引擎/隧道，下次访问用新配置重建；同步清健康位（新配置视作全新开始）
        self.pool.dispose_connection(project, connection)
        self.redis_pool.dispose_connection(project, connection)
        self.health.force_clear(project, connection)
        rec = AuditRecord(project=project, connection=connection, tool=tool, status="ok",
                          agent=caller.agent, session_id=caller.session_id, detail=detail)
        self.store.record(rec)

    # ---------- 后台维护（serve 时启动）----------

    def start_housekeeping(
        self,
        retention_days: int = DEFAULT_RETENTION_DAYS,
        interval_s: int = HOUSEKEEPING_INTERVAL_S,
    ) -> None:
        """周期任务：空闲引擎/隧道回收 + 审计与终态审批单按保留期清理。"""
        if self._housekeeping_stop is not None:
            return
        stop = threading.Event()
        self._housekeeping_stop = stop

        def _loop() -> None:
            while not stop.wait(interval_s):
                self.housekeep_once(retention_days)

        threading.Thread(target=_loop, name="dbm-housekeeping", daemon=True).start()

    def housekeep_once(self, retention_days: int = DEFAULT_RETENTION_DAYS) -> dict:
        """执行一轮维护，返回统计（供测试与日志）。单项失败不影响其他项。"""
        from .inbox import DEFAULT_RETENTION_DAYS as INBOX_RETENTION
        stats = {"engines_reaped": 0, "redis_reaped": 0, "audit_purged": 0,
                 "changes_purged": 0, "notifications_purged": 0,
                 "workflow_runs_purged": 0, "exports_purged": 0}
        for key, fn in (
            ("engines_reaped", self.pool.reap_idle),
            ("redis_reaped", self.redis_pool.reap_idle),
            ("audit_purged", lambda: self.store.purge_old(retention_days)),
            ("changes_purged",
             (lambda: self.approvals.purge_old(retention_days)) if self.approvals else (lambda: 0)),
            # 通知短保留（7 天）：审批提醒/exhausted 告警不必久存
            ("notifications_purged",
             (lambda: self.inbox.purge_old(INBOX_RETENTION)) if self.inbox else (lambda: 0)),
            # workflow_run：30 天保留；同时清理 xlsx 产物目录
            ("workflow_runs_purged", self._purge_workflow_runs),
            # MCP 临时导出：固定一小时 TTL，与审计保留期无关
            ("exports_purged", self.purge_mcp_exports),
        ):
            try:
                stats[key] = fn()
            except Exception:
                logger.exception("housekeeping %s 失败", key)
        if any(stats.values()):
            logger.info("housekeeping: %s", stats)
        return stats

    def _purge_workflow_runs(self, days: int = 30) -> int:
        """清 30 天前的 workflow_run 记录 + 顺手删对应 xlsx 目录。"""
        if self.runs is None:
            return 0
        paths = self.runs.purge_older_than(days)
        if not paths or not self.data_dir:
            return len(paths)
        import shutil
        from pathlib import Path
        for rel in paths:
            try:
                full = Path(self.data_dir) / rel
                if full.parent.is_dir():
                    shutil.rmtree(full.parent, ignore_errors=True)
            except Exception:  # noqa: BLE001
                logger.exception("清理 xlsx 目录失败：%s", rel)
        return len(paths)

    # ---------- 调度：CRUD + tick 循环 ----------

    def workflow_schedule_upsert(self, name: str, cron_type: str, cron_value: str,
                                 enabled: bool = True, notify_on: str = "failure",
                                 attach_kinds: list[str] | None = None) -> dict:
        """增/改一条调度配置。校验 cron 语法和 notify_on/attach_kinds 值域。"""
        if self.schedules is None:
            raise RuntimeError("调度存储未初始化（需 serve 模式运行）")
        # workflow 存在性检查（调度只能挂在已有 workflow 上）
        if self.workflows is None or not any(w.name == name for w in self.workflows.list()):
            raise ValueError(f"workflow {name!r} 不存在")
        return self.schedules.upsert(name, cron_type, cron_value, enabled=enabled,
                                     notify_on=notify_on, attach_kinds=attach_kinds)

    def workflow_schedule_get(self, name: str) -> dict | None:
        if self.schedules is None:
            return None
        return self.schedules.get(name)

    def workflow_schedule_delete(self, name: str) -> None:
        if self.schedules is None:
            return
        self.schedules.delete(name)

    def workflow_schedule_list(self) -> list[dict]:
        if self.schedules is None:
            return []
        return self.schedules.list()

    def workflow_schedules_enriched(self) -> list[dict]:
        """定时任务全局列表：每条 schedule 附上 workflow 是否存在、是否有正在跑的实例。

        供 /admin/workflows/schedules 管理页用；纯只读，不改任何状态。
        """
        if self.schedules is None:
            return []
        scheds = self.schedules.list()
        wf_names = set()
        if self.workflows is not None:
            wf_names = {w.name for w in self.workflows.list()}
        out = []
        for s in scheds:
            name = s.get("name")
            running = False
            if self.runs is not None and name:
                running = bool(self.runs.running_for(name))
            out.append({**s, "workflow_exists": name in wf_names, "running": running})
        return out

    def workflow_schedule_trigger_now(self, name: str) -> None:
        """手动立即触发一次调度（后台线程跑，不阻塞调用方）。

        走完全跟 cron 到点相同的链路：入 workflow_run → 后台跑 → 通知。
        与用户在流程详情页 ▶ 运行的区别：那个是前台同步、不入 workflow_run 表、不发通知；
        这个是把「计划本来会自动跑的一次」提前到现在。
        """
        if self.workflows is None or self.workflows.get(name) is None:
            raise ValueError(f"workflow 不存在：{name!r}")
        if self.schedules is None or self.schedules.get(name) is None:
            raise ValueError(f"调度配置不存在：{name!r}")
        threading.Thread(
            target=self._run_scheduled, args=(name,),
            daemon=True, name=f"dbm-wf-run-{name}").start()

    # ---------- 运行历史（供详情页 & 通知 deeplink 用）----------

    def workflow_runs_list(self, name: str, limit: int = 50) -> list[dict]:
        if self.runs is None:
            return []
        return self.runs.list_by_name(name, limit)

    def workflow_run_get(self, run_id: int) -> dict | None:
        if self.runs is None:
            return None
        return self.runs.get(run_id)

    def workflow_running_list(self, triggered_by: str | None = "schedule") -> list[dict]:
        """当前正在执行的 workflow 运行列表；默认只列调度触发的（手动/agent 触发的用户在前端等结果，不列）。"""
        if self.runs is None:
            return []
        from datetime import UTC, datetime
        rows = self.runs.list_running(triggered_by)
        now_ts = datetime.now(UTC).timestamp()
        out = []
        for r in rows:
            elapsed = None
            started = r.get("started_at")
            if started:
                try:
                    elapsed = int(now_ts - datetime.fromisoformat(started).timestamp())
                except ValueError:
                    elapsed = None
            out.append({"id": r["id"], "name": r["name"],
                        "triggered_by": r["triggered_by"],
                        "started_at": started, "elapsed_s": elapsed})
        return out

    # ---------- 调度触发的一次执行 ----------

    def _run_scheduled(self, name: str) -> None:
        """一次调度触发：新建 workflow_run → 后台跑 → 更新状态 → 按 notify_on 发通知。

        同名 workflow 上一次尚未完成 → 跳过本次（防堆积；对齐"安静即正常"红线）。
        任何异常吞掉（scheduler tick 不应因单个 workflow 失败而挂）。
        """
        if self.runs is None or self.workflows is None:
            return
        try:
            if self.runs.running_for(name):
                logger.warning("scheduler: workflow %s 上次未完成，跳过本次调度", name)
                return
            sched = self.schedules.get(name) if self.schedules is not None else None
            run_id = self.runs.start(name, triggered_by="schedule")
            # 走既有 workflow_run（此处 caller 用 scheduler；审计记录里区分）
            sched_caller = CallerInfo(agent="scheduler", session_id=f"sched-{run_id}")
            try:
                out = self.workflow_run(name, sched_caller)
                ok = out.get("ok") is True
                # 输出预览：只留可控大小（保存整个 rows 的话表膨胀），前 100 行
                op = out.get("output") or None
                op_saved: dict = {}
                if op:
                    op_saved = {"columns": op.get("columns") or [],
                                "rows": (op.get("rows") or [])[:100],
                                "row_count": op.get("row_count", 0)}
                # xlsx 产物：attach_kinds 含 xlsx_link 才生成
                xlsx_path: str | None = None
                attach_kinds = (sched or {}).get("attach_kinds") or []
                if ok and op and "xlsx_link" in attach_kinds and self.data_dir:
                    xlsx_path = self._save_run_xlsx(run_id, op)
                self.runs.finish(run_id, "ok" if ok else "failed",
                                 steps=out.get("steps") or [],
                                 output_preview=op_saved,
                                 error="" if ok else (
                                     (out.get("steps") or [{}])[-1].get("error", "") or "运行失败"),
                                 xlsx_path=xlsx_path)
            except Exception as e:  # noqa: BLE001
                self.runs.finish(run_id, "failed", error=f"{type(e).__name__}: {e}")
                ok = False
            # 更新 schedule 元数据
            if self.schedules is not None:
                try:
                    self.schedules.mark_ran(name, "ok" if ok else "failed")
                except Exception:  # noqa: BLE001
                    logger.exception("mark_ran 失败：%s", name)
            # 通知
            self._notify_scheduled_run(name, run_id, sched)
        except Exception:  # noqa: BLE001
            logger.exception("_run_scheduled 未预期失败：%s", name)

    def _save_run_xlsx(self, run_id: int, output: dict) -> str | None:
        """把 workflow_run 的输出落成 xlsx，返回相对 data_dir 的路径（或 None）。"""
        from pathlib import Path

        from . import export
        try:
            cols = output.get("columns") or []
            rows = output.get("rows") or []
            if not cols:
                return None
            rel = f"workflow_runs/{run_id}/output.xlsx"
            full = Path(self.data_dir) / rel
            full.parent.mkdir(parents=True, exist_ok=True)
            data = export.to_xlsx(cols, rows)
            full.write_bytes(data)
            return rel
        except Exception:  # noqa: BLE001
            logger.exception("生成 xlsx 失败：run_id=%s", run_id)
            return None

    def _notify_scheduled_run(self, name: str, run_id: int, sched: dict | None) -> None:
        """按 notify_on 决定是否发通知；富通知走 render_workflow_notification。"""
        if not sched:
            return
        notify_on = sched.get("notify_on") or "failure"
        if notify_on == "none":
            return
        run = self.runs.get(run_id) if self.runs is not None else None
        if not run:
            return
        ok = run.get("status") == "ok"
        if notify_on == "success" and not ok:
            return
        if notify_on == "failure" and ok:
            return
        try:
            from .notify import render_workflow_notification
            settings = self.get_settings() if self.settings is not None else {}
            admin_base_url = (settings.get("admin_base_url") or "").strip()
            download_path = None
            if run.get("xlsx_path"):
                download_path = f"/admin/workflows/runs/{run_id}/download/output.xlsx"
            payload = render_workflow_notification(
                name, run, sched.get("attach_kinds") or ["summary"],
                admin_base_url=admin_base_url, download_path=download_path)
            self.notifier.send(payload["title"], payload["body"], meta=payload["meta"])
        except Exception:  # noqa: BLE001
            logger.exception("workflow 通知发送失败：run_id=%s", run_id)

    def start_scheduler(self, interval_s: int = 30) -> None:
        """启动调度器 daemon 线程。tick 粒度 30s（下拉最小 1min，抖动 ≤30s）。

        安静即正常红线：__init__ 默认不启动（测试路径永不真跑），只 _cmd_serve 显式调。
        启动时先扫一遍 workflow_run 把 1h 前还挂 running 的记录标 failed（防重启阻塞）。
        """
        if self._scheduler_stop is not None or self.schedules is None or self.runs is None:
            return
        try:
            swept = self.runs.sweep_stale_running(older_than_hours=1)
            if swept:
                logger.info("scheduler: sweep %d stale running", swept)
        except Exception:  # noqa: BLE001
            logger.exception("scheduler: sweep stale 失败")
        stop = threading.Event()
        self._scheduler_stop = stop

        def _loop() -> None:
            while not stop.wait(interval_s):
                try:
                    self._scheduler_tick()
                except Exception:  # noqa: BLE001
                    logger.exception("scheduler tick 失败")

        threading.Thread(target=_loop, daemon=True, name="dbm-scheduler").start()

    def _scheduler_tick(self) -> None:
        """遍历 enabled 调度 → cron_matches(now) → 触发。同分钟内每个 workflow 只跑一次。"""
        if self.schedules is None:
            return
        from datetime import datetime

        from .workflows import cron_from_dropdown, cron_matches
        now = datetime.now()
        stamp = now.strftime("%Y-%m-%d %H:%M")
        for sched in self.schedules.list_enabled():
            name = sched["name"]
            key = (name, stamp)
            if key in self._sched_ticked_minute:
                continue
            try:
                cron_expr = cron_from_dropdown(sched["cron_type"], sched["cron_value"])
            except ValueError:
                logger.exception("scheduler: 非法 cron %s / %s / %s",
                                 name, sched["cron_type"], sched["cron_value"])
                continue
            if not cron_matches(cron_expr, now):
                continue
            self._sched_ticked_minute.add(key)
            # 清理旧分钟条目（防内存增长）
            if len(self._sched_ticked_minute) > 10000:
                self._sched_ticked_minute.clear()
            # 后台线程跑，别阻塞 tick 循环
            threading.Thread(
                target=self._run_scheduled, args=(name,),
                daemon=True, name=f"dbm-wf-run-{name}").start()

    def close(self) -> None:
        if self._housekeeping_stop is not None:
            self._housekeeping_stop.set()
            self._housekeeping_stop = None
        if self._scheduler_stop is not None:
            self._scheduler_stop.set()
            self._scheduler_stop = None
        self.health.stop()
        self.pool.dispose()
        self.redis_pool.dispose()
        self.store.close()
        if self.approvals is not None:
            self.approvals.close()
        if self.metadata is not None:
            self.metadata.close()
        if self.snippets is not None:
            self.snippets.close()
        if self.settings is not None:
            self.settings.close()
        if self.inbox is not None:
            self.inbox.close()
        if self.schedules is not None:
            self.schedules.close()
        if self.runs is not None:
            self.runs.close()
