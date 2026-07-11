"""导出序列化纯函数测试：CSV / JSON / Markdown / Excel。"""

import json

import pytest

from dbmcp.export import (
    ExportError,
    export_result,
    to_csv,
    to_json,
    to_markdown,
)

COLUMNS = ["id", "name", "blob"]
ROWS = [
    [1, "alice", None],
    [2, "b|ob\nnext", {"__bytes_base64__": "QUJD"}],
]


def test_csv_has_bom_and_header_and_bytes_cell():
    out = to_csv(COLUMNS, ROWS)
    assert out.startswith(b"\xef\xbb\xbf")  # UTF-8 BOM，Excel 中文不乱码
    text = out.decode("utf-8-sig")
    assert text.splitlines()[0] == "id,name,blob"
    assert "base64:QUJD" in text
    # None → 空串
    assert ",alice," in text


def test_json_roundtrip_and_null():
    out = to_json(COLUMNS, ROWS)
    data = json.loads(out)
    assert data[0] == {"id": 1, "name": "alice", "blob": None}
    assert data[1]["blob"] == {"__bytes_base64__": "QUJD"}


def test_markdown_escapes_pipe_and_newline():
    out = to_markdown(COLUMNS, ROWS).decode("utf-8")
    lines = out.strip().splitlines()
    assert lines[0] == "| id | name | blob |"
    assert lines[1] == "| --- | --- | --- |"
    # 管道被转义、换行被压平，避免破坏表格结构
    assert r"b\|ob next" in out
    assert "\n" not in lines[3] if len(lines) > 3 else True


def test_export_result_returns_media_type_and_ext():
    data, media_type, ext = export_result(COLUMNS, ROWS, "csv")
    assert ext == "csv" and "text/csv" in media_type and data.startswith(b"\xef\xbb\xbf")
    data, media_type, ext = export_result(COLUMNS, ROWS, "json")
    assert ext == "json" and "json" in media_type


def test_export_result_rejects_unknown_format():
    with pytest.raises(ExportError, match="不支持"):
        export_result(COLUMNS, ROWS, "parquet")


def test_xlsx_produces_valid_workbook():
    pytest.importorskip("openpyxl")
    import io

    from openpyxl import load_workbook

    from dbmcp.export import to_xlsx

    data = to_xlsx(COLUMNS, ROWS)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["id", "name", "blob"]
    assert ws[2][0].value == 1 and ws[2][1].value == "alice"
    # 二进制 dict 落格转文本
    assert ws[3][2].value == "base64:QUJD"
